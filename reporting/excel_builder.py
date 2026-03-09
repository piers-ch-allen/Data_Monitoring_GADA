import pandas as pd
from openpyxl.styles import PatternFill
import math
from openpyxl.formatting.rule import FormulaRule
from reporting.comparison import compare_previous_month
from openpyxl.drawing.image import Image as XLImage
from reporting.patient_tracking import plot_df, plot_df_center, plot_df_follow_ups, calc_follow_up_avg
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def format_worksheet_with_max_width(
    worksheet,
    max_width=40,
    base_row_height=15
):
    for col in worksheet.columns:
        col_letter = get_column_letter(col[0].column)

        max_length = 0

        for cell in col:
            if cell.value:
                cell.alignment = Alignment(wrap_text=True)

                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length

        adjusted_width = min(max_length + 2, max_width)

        worksheet.column_dimensions[col_letter].width = adjusted_width

    # Second: Adjust row heights based on wrapped text
    for row in worksheet.iter_rows():
        max_height = base_row_height

        for cell in row:
            if cell.value:
                col_letter = get_column_letter(cell.column)
                col_width = worksheet.column_dimensions[col_letter].width

                text_length = len(str(cell.value))

                # Estimate wrapped lines
                estimated_lines = math.ceil(text_length / col_width)

                cell_height = estimated_lines * base_row_height

                max_height = max(max_height, cell_height)

        worksheet.row_dimensions[row[0].row].height = max_height

def summary_builder(
    missed_visit_df,
    due_visit_df,
    issues_df_req,
    issues_df_opt,
    files_info,
    visit_group
):
    # ---- Basic validation ----
    required_dfs = {
        "missed_visit_df": missed_visit_df,
        "due_visit_df": due_visit_df,
        "issues_df_req": issues_df_req,
        "issues_df_opt": issues_df_opt
    }

    # ---- Unique patient calculations ----
    unique_req = (
        issues_df_req["ID"].dropna().nunique()
        if "ID" in issues_df_req.columns else 0
    )

    unique_opt = (
        issues_df_opt["ID"].dropna().nunique()
        if "ID" in issues_df_opt.columns else 0
    )

    combined_unique = pd.concat(
        [
            issues_df_req["ID"].dropna() if "ID" in issues_df_req.columns else pd.Series(dtype=str),
            issues_df_opt["ID"].dropna() if "ID" in issues_df_opt.columns else pd.Series(dtype=str),
        ]
    ).nunique()

    average_follow_up = calc_follow_up_avg(visit_group)

    # ---- Build summary ----
    summary_df = pd.DataFrame({
        "Report Section": [
            "Recruitment Information",
            "Total Recruited: Doctor",
            "Total Follow-Ups Visit 1",
            "Total Follow-Ups Visit 2",
            "Total Follow-Ups Visit 3",
            "Patient Visits Overdue",
            "Patient Visits Due",
            " ",
            "Recorded Issues",
            "Required Data Points Issues Recorded",
            "Optional Data Points Issues Recorded",
            "Unique Patient Issues (Required)",
            "Unique Patient Issues (Optional)",
            "Unique Patient Issues (All)",
            " "
        ],
        "Value": [
            "",
            files_info.get("Doctor - Visit 1", 0),
            files_info.get("Doctor - Visit 2", 0),
            files_info.get("Doctor - Visit 3", 0),
            files_info.get("Doctor - Visit 4", 0),
            len(missed_visit_df) if missed_visit_df is not None else 0,
            len(due_visit_df) if due_visit_df is not None else 0,
            "",
            "",
            len(issues_df_req) if issues_df_req is not None else 0,
            len(issues_df_opt) if issues_df_opt is not None else 0,
            unique_req,
            unique_opt,
            combined_unique,
            " "
        ]
    })
    return summary_df, average_follow_up

def sheet_comparer(
    sheets,
    writer
):
    highlight_fill = PatternFill(start_color="FFF2CC",
                                 end_color="FFF2CC",
                                 fill_type="solid")
    for sheet_name, df in sheets.items():
        worksheet = writer.sheets[sheet_name]
        if worksheet.max_row <= 1:
            continue
        col_index = df.columns.get_loc("Exists_Last_Month") + 1
        col_letter = get_column_letter(col_index)

        max_row = worksheet.max_row
        max_col = worksheet.max_column
        last_col_letter = get_column_letter(max_col)

        formula = (
            f'=AND(${col_letter}2=TRUE,'
            f'ISERROR(SEARCH("doctor",$A2)),'
            f'ISERROR(SEARCH("patient",$A2)),'
            f'$A2<>" ")'
        )

        rule = FormulaRule(
            formula=[formula],
            fill=highlight_fill
        )

        worksheet.conditional_formatting.add(
            f"A2:{last_col_letter}{max_row}",
            rule
        )

def summary_graph_builder(
    visit_group,
    summary_df,
    missed_df,
    due_df,
    visit_df,
    writer
):
    # Gather the graphs

    import matplotlib
    matplotlib.use("Agg")
    matplotlib.rcParams['font.family'] = 'DejaVu Sans'
    import matplotlib.pyplot as plt

    tmp_total = plot_df(visit_group[0], plt)
    tmp_centre = plot_df_center(visit_group[0], plt)
    tmp_followup = plot_df_follow_ups(visit_group, missed_df, due_df, plt)
    img1 = XLImage(tmp_total.name)
    img2 = XLImage(tmp_centre.name)
    img3 = XLImage(tmp_followup.name)
    # Resize images for Excel fit
    img1.width = 620
    img1.height = 360
    img2.width = 800
    img2.height = 500
    img3.width = 800
    img3.height = 500

    # Export the graphs to an Excel sheet
    workbook = writer.book
    sheet_name = "Summary of Recruitment"
    worksheet = workbook.create_sheet(sheet_name)
    # Write first dataframe
    summary_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Calculate the next row after summary_df
    start_row = len(summary_df) + 2  # +2 adds a blank row between tables

    # Write second dataframe below the first
    visit_df.to_excel(
        writer,
        sheet_name=sheet_name,
        startrow=start_row,
        index=False
    )
    worksheet.add_image(img1, "F25")
    worksheet.add_image(img2, "F1")
    worksheet.add_image(img3, "R1")

def build_excel_report(
    missed_df,
    due_df,
    req,
    opt,
    prev_file,
    visit_group,
    files_info,
    report_path
):
    # ---- Create the required dataFrames ----
    issues_req_df = pd.DataFrame(req)
    issues_opt_df = pd.DataFrame(opt)
    summary_df, visit_df = summary_builder(missed_df, due_df, issues_req_df, issues_opt_df, files_info, visit_group)

    # Check for values present in the previous month
    missed_df, due_df, issues_req_df, issues_opt_df = compare_previous_month(missed_df,
                                                                             due_df,
                                                                             issues_req_df,
                                                                             issues_opt_df,
                                                                             prev_file)
    sheets = {
        "Patient Visit Missed": missed_df,
        "Patient Visit Due": due_df,
        "Required Data Point Issues": issues_req_df,
        "Optional Data Point Issues": issues_opt_df,
    }

    # ---- Write Report ----
    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        # Write all the data to the Excel file
        summary_graph_builder(visit_group, summary_df, missed_df, due_df, visit_df, writer)
        for sheet_name_val, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name_val, index=False)
        for sheet in writer.sheets.values():
            format_worksheet_with_max_width(sheet, 60)
        sheet_comparer(sheets, writer)



